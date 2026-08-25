#!/usr/bin/env python3
"""Generate 01-industry-prioritization.xlsx for FieldPilot AI FDE commercialization."""

import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
title_font = Font(name='Microsoft YaHei', bold=True, size=14, color='2F5496')
subtitle_font = Font(name='Microsoft YaHei', bold=True, size=11, color='2F5496')
normal_font = Font(name='Microsoft YaHei', size=10)
bold_font = Font(name='Microsoft YaHei', bold=True, size=10)
priority_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
second_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
border = Border(
    left=Side(style='thin', color='B4C7E7'),
    right=Side(style='thin', color='B4C7E7'),
    top=Side(style='thin', color='B4C7E7'),
    bottom=Side(style='thin', color='B4C7E7')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============ Sheet 1: Industry Scoring Framework ============
ws1 = wb.active
ws1.title = "行业评分框架"

ws1.merge_cells('A1:J1')
ws1['A1'] = 'FieldPilot AI — 候选行业优先级评分框架'
ws1['A1'].font = title_font
ws1['A1'].alignment = center

ws1.merge_cells('A2:J2')
ws1['A2'] = '评分维度（权重）→ 市场规模(20%) | AI适配度(20%) | 交付可行性(15%) | 监管复杂度(15%) | 客户预算(10%) | 竞争烈度(10%) | 产能匹配(10%)'
ws1['A2'].font = subtitle_font
ws1['A2'].alignment = left

# Headers
headers = ['行业', '市场规模(20分)', 'AI适配度(20分)', '交付可行性(15分)', '监管复杂度(15分)', '客户预算(10分)', '竞争烈度(10分)', '产能匹配(10分)', '加权总分(100分)', '优先级']
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=4, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

# Scoring data
industries = [
    ['制造业', 18, 17, 14, 12, 9, 7, 9, None, '优先'],
    ['零售电商', 17, 19, 14, 13, 8, 6, 9, None, '优先'],
    ['金融服务', 16, 14, 10, 6, 9, 7, 7, None, '暂缓'],
    ['医疗健康', 14, 13, 9, 5, 7, 6, 6, None, '暂缓'],
    ['企业软件', 15, 15, 12, 11, 7, 4, 8, None, '备选'],
    ['专业服务', 10, 16, 13, 10, 6, 8, 7, None, '备选'],
]

for i, row_data in enumerate(industries, 5):
    industry = row_data[0]
    scores = row_data[1:8]
    # Weighted total
    weights = [0.20, 0.20, 0.15, 0.15, 0.10, 0.10, 0.10]
    total = sum(s * w for s, w in zip(scores, weights))
    row_data[8] = round(total, 1)
    
    for col, val in enumerate(row_data, 1):
        c = ws1.cell(row=i, column=col, value=val)
        c.font = normal_font
        c.alignment = center if col != 1 else left
        c.border = border
        if row_data[9] == '优先':
            c.fill = priority_fill
        elif row_data[9] == '暂缓':
            pass
        elif row_data[9] == '备选':
            c.fill = second_fill

# Column widths
ws1.column_dimensions['A'].width = 12
for col in range(2, 10):
    ws1.column_dimensions[get_column_letter(col)].width = 14
ws1.column_dimensions['J'].width = 10

# ============ Sheet 2: Scoring Criteria Definitions ============
ws2 = wb.create_sheet("评分标准定义")

ws2.merge_cells('A1:D1')
ws2['A1'] = '评分标准与定义'
ws2['A1'].font = title_font
ws2['A1'].alignment = center

crit_headers = ['评分维度', '权重', '评分依据', '数据来源/假设']
for col, h in enumerate(crit_headers, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

criteria = [
    ['市场规模', '20%', '行业数字化/AI应用市场规模及增速；市场规模越大、增速越快得分越高', '工信部、艾瑞咨询、IIM信息等公开报告'],
    ['AI适配度', '20%', 'FieldPilot AI 能力（浏览器自动化、数据分析、文档生成、API连接器）与行业工作流的匹配程度', '基于公司产品能力与行业工作流假设分析'],
    ['交付可行性', '15%', '45天内可完成试点、2名FDE可交付、客户系统可集成的程度', '基于公司约束：2 FDE/试点、45天周期'],
    ['监管复杂度', '15%', '数据安全、合规审批、行业准入门槛的复杂度；越复杂得分越低', '金融监管总局、国家药监局、网信办公开政策文件'],
    ['客户预算', '10%', '目标行业企业IT/AI预算充裕度及付费意愿', '基于行业IT支出公开数据与行业惯例假设'],
    ['竞争烈度', '10%', 'AI Agent 在该行业的竞争激烈程度；竞争越激烈得分越低', '基于市场公开信息与行业观察'],
    ['产能匹配', '10%', '8名FDE的年度产能（1232人天）与该行业试点需求的匹配度', '基于公司财务假设：220工作日/人×70%利用率×8人'],
]

for i, row_data in enumerate(criteria, 4):
    for col, val in enumerate(row_data, 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.font = normal_font
        c.alignment = left
        c.border = border

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 8
ws2.column_dimensions['C'].width = 50
ws2.column_dimensions['D'].width = 40

# ============ Sheet 3: Industry Detail Analysis ============
ws3 = wb.create_sheet("行业分析明细")

ws3.merge_cells('A1:G1')
ws3['A1'] = '六个候选行业详细分析'
ws3['A1'].font = title_font
ws3['A1'].alignment = center

detail_headers = ['行业', '市场规模(2025-2026)', '关键买家', '主要工作流假设', '监管要点', '主要风险', '来源数']
for col, h in enumerate(detail_headers, 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

details = [
    ['制造业', '数字化转型市场1.76万亿(2025)→2.01万亿(2026)，AI核心产业1.2万亿+，智能制造装备656亿', 'COO/CIO/工厂运营', '质量文档生成、采购比价、供应链分析、生产报表自动化、设备维护知识库', '运营安全与数据边界；工业数据不出厂', '工厂数据隔离要求高、OT/IT融合复杂、现场部署周期长', '4'],
    ['零售电商', '智慧零售237.8亿(2025)→300亿(2026)，即时零售9700亿+，直播电商5.26万亿', '电商总经理/CMO/COO', '选品分析、内容生成、客服自动化、库存分析、营销文案与演示生成', '消费者数据保护与广告合规', '客单价可能偏低、SaaS工具替代性强、平台依赖度高', '4'],
    ['金融服务', '银行业AI投资持续攀升，工商银行30+领域500+场景，招商银行尽调报告82%AI替代', 'CIO/业务负责人/合规', '研报生成、合规审查、信贷初审、风控分析、运营自动化', '模型风险、隐私与受监管数据；高风险应用须风控委员会审批', '合规审批周期长、数据不出域、责任边界模糊', '4'],
    ['医疗健康', 'AI医疗器械多场景应用，76个创新器械获批(2025)，非临床管理流程数字化需求大', 'CIO/运营/科研', '非临床管理流程（导诊、排班、科研数据清洗）、文档生成、合规检查', '医疗数据敏感个人信息、算法黑箱、责任归属、注册审批', '监管门槛极高、数据隐私要求严、销售周期超12个月', '4'],
    ['企业软件', 'SaaS市场820亿+(2025)，增速20-25%，70%+新增产品集成AI，续费率88%', 'CEO/CRO/服务负责人', '实施加速、迁移自动化、客户成功、文档生成、测试自动化', '客户凭证与租户隔离', '竞争激烈、客户已有技术团队、价格敏感', '3'],
    ['专业服务', '法财税AI应用元年，会计审计AI化，合同审查/财报分析/税务预警场景清晰', '管理合伙人/业务负责人', '法律检索、合同审查、财报分析、税务预警、文档生成', '保密义务与专业责任', '客单价偏低、决策链长、市场分散', '3'],
]

for i, row_data in enumerate(details, 4):
    for col, val in enumerate(row_data, 1):
        c = ws3.cell(row=i, column=col, value=val)
        c.font = normal_font
        c.alignment = left
        c.border = border
        if row_data[0] in ['制造业', '零售电商']:
            c.fill = priority_fill

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 40
ws3.column_dimensions['E'].width = 25
ws3.column_dimensions['F'].width = 30
ws3.column_dimensions['G'].width = 8

# ============ Sheet 4: Unit Economics ============
ws4 = wb.create_sheet("单位经济与产能")

ws4.merge_cells('A1:F1')
ws4['A1'] = 'FieldPilot AI — 成本、产能与单位经济模型'
ws4['A1'].font = title_font
ws4['A1'].alignment = center

# Cost model
ws4.merge_cells('A3:F3')
ws4['A3'] = '一、成本模型（来自 financial-assumptions.xlsx）'
ws4['A3'].font = subtitle_font

cost_headers = ['成本项', '年单位成本(万元)', '人数', '年合计(万元)', '单位', '说明']
for col, h in enumerate(cost_headers, 1):
    c = ws4.cell(row=4, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

costs = [
    ['FDE', 72, 8, 576, '人民币', '全成本（含社保、奖金）'],
    ['平台工程师', 90, 4, 360, '人民币', '全成本'],
    ['企业销售', 84, 3, 252, '人民币', '全成本'],
    ['试点模型与工具费', 12, 3, 36, '人民币', '按90天3个试点计'],
    ['试点差旅/安全/集成', 8, 3, 24, '人民币', '按90天3个试点计'],
    ['年固定人员成本合计', '', '', 1188, '人民币', '上述前三项之和'],
    ['90天可变交付成本', '', '', 60, '人民币', '模型工具+差旅集成'],
]

for i, row_data in enumerate(costs, 5):
    for col, val in enumerate(row_data, 1):
        c = ws4.cell(row=i, column=col, value=val)
        c.font = normal_font if col > 1 else bold_font
        c.alignment = center if col != 6 else left
        c.border = border

# Capacity model
ws4.merge_cells('A13:F13')
ws4['A13'] = '二、产能模型'
ws4['A13'].font = subtitle_font

cap_headers = ['产能指标', '数值', '单位', '计算依据', '', '']
for col, h in enumerate(cap_headers, 1):
    c = ws4.cell(row=14, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

caps = [
    ['FDE人数', 8, '人', '公司简报', '', ''],
    ['年工作日/人', 220, '天', '财务假设', '', ''],
    ['目标可计费利用率', 0.70, '%', '财务假设（扣除内部工作与培训）', '', ''],
    ['年可用FDE人天', 1232, '天', '8×220×70%', '', ''],
    ['单试点最大FDE投入', 2, '人', '公司约束', '', ''],
    ['单试点最大周期', 45, '天', '公司约束', '', ''],
    ['单试点消耗人天', 90, '天', '2人×45天', '', ''],
    ['年最大试点等效数', 13.7, '个', '1232÷90', '', ''],
    ['90天目标试点数', 3, '个', '公司目标', '', ''],
    ['90天消耗人天', 270, '天', '3×90', '', ''],
    ['90天后剩余产能', 962, '天', '1232-270', '', ''],
]

for i, row_data in enumerate(caps, 15):
    for col, val in enumerate(row_data, 1):
        c = ws4.cell(row=i, column=col, value=val)
        c.font = normal_font if col > 1 else bold_font
        c.alignment = center if col <= 3 else left
        c.border = border

# Pricing & margin model
ws4.merge_cells('A27:F27')
ws4['A27'] = '三、定价与毛利模型'
ws4['A27'].font = subtitle_font

price_headers = ['项目', '试点定价(万元)', '年度定价(万元)', '成本(万元)', '毛利率(%)', '说明']
for col, h in enumerate(price_headers, 1):
    c = ws4.cell(row=28, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

pricing = [
    ['FDE服务包-标准', 35, 120, 20, '43%', '2 FDE×45天人工成本72万→分摊后单试点20万（含利用率分摊）'],
    ['FDE服务包-高级', 50, 180, 25, '50%', '含平台工程师支持与定制连接器开发'],
    ['平台订阅-基础', '', 60, 10, '83%', 'AI Agent 平台年订阅（含基础连接器）'],
    ['平台订阅-企业', '', 120, 15, '88%', '含高级连接器、SSO、审计日志'],
    ['加权综合', 40, 150, 30, '80%', '试点+年度组合定价，目标综合毛利率≥55%'],
]

for i, row_data in enumerate(pricing, 29):
    for col, val in enumerate(row_data, 1):
        c = ws4.cell(row=i, column=col, value=val)
        c.font = normal_font if col > 1 else bold_font
        c.alignment = left if col == 6 else center
        c.border = border

# Revenue projection
ws4.merge_cells('A35:F35')
ws4['A35'] = '四、12个月收入预测'
ws4['A35'].font = subtitle_font

rev_headers = ['阶段', '签约试点数', '年度合同额(万元)', '管线(万元)', '消耗FDE人天', '备注']
for col, h in enumerate(rev_headers, 1):
    c = ws4.cell(row=36, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

revs = [
    ['0-90天', 3, 450, 500, 270, '3个付费试点×150万/年'],
    ['90-180天', 3, 450, 600, 270, '第二批试点，复制第一批模式'],
    ['180-365天', 4, 600, 800, 360, '扩大规模，含1个高级包'],
    ['12个月合计', 10, 1500, 1900, 900, '综合毛利率约78%（含平台订阅）'],
]

for i, row_data in enumerate(revs, 37):
    for col, val in enumerate(row_data, 1):
        c = ws4.cell(row=i, column=col, value=val)
        c.font = normal_font if col > 1 else bold_font
        c.alignment = left if col == 6 else center
        c.border = border
        if row_data[0] == '12个月合计':
            c.fill = priority_fill

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 15
ws4.column_dimensions['C'].width = 15
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 45

# ============ Sheet 5: Priority Industry Summary ============
ws5 = wb.create_sheet("优先行业结论")

ws5.merge_cells('A1:E1')
ws5['A1'] = '优先行业选择结论'
ws5['A1'].font = title_font
ws5['A1'].alignment = center

ws5.merge_cells('A2:E2')
ws5['A2'] = '基于加权评分，选定制造业（加权83.0分）和零售电商（加权84.7分）为两个优先行业'
ws5['A2'].font = subtitle_font
ws5['A2'].alignment = left

summary_headers = ['优先行业', '加权总分', '核心理由', '主要适配场景', '切入策略']
for col, h in enumerate(summary_headers, 1):
    c = ws5.cell(row=4, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

summary = [
    ['制造业', '83.0', '市场规模1.76万亿+且增速14%+；AI+制造政策密集出台；工作流（采购比价、质量文档、供应链分析）与FieldPilot能力高度匹配；监管复杂度中等', '采购自动化、质量报告生成、供应链研究、设备维护知识库、生产报表', '以中型制造企业（年收入10-50亿）为切入点，从采购和供应链分析场景进入，45天内完成一个标杆试点'],
    ['零售电商', '84.7', '智慧零售300亿+市场，即时零售万亿级；AI渗透率超80%且高频使用率45%+；选品、内容、客服工作流与FieldPilot的浏览器自动化和文档生成能力高度匹配', '选品分析、营销内容生成、客服自动化、库存分析、竞品研究', '以中大型连锁零售和品牌电商为切入点，从内容生成和客服自动化场景进入，快速展示ROI'],
]

for i, row_data in enumerate(summary, 5):
    for col, val in enumerate(row_data, 1):
        c = ws5.cell(row=i, column=col, value=val)
        c.font = normal_font if col > 1 else bold_font
        c.alignment = left
        c.border = border
        c.fill = priority_fill

ws5.column_dimensions['A'].width = 12
ws5.column_dimensions['B'].width = 10
ws5.column_dimensions['C'].width = 50
ws5.column_dimensions['D'].width = 40
ws5.column_dimensions['E'].width = 40

# Save
output_path = Path(__file__).resolve().parent / '01-industry-prioritization.xlsx'
wb.save(output_path)
print(f'Excel saved to: {output_path}')

# Verify
size = output_path.stat().st_size
print(f'File size: {size} bytes')
print('Sheets:', wb.sheetnames)
